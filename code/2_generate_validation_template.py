"""
ARCG Experiment  --  Step 2: Generate Human Validation Template
================================================================
Reads data/paraphrases.json and produces two Excel workbooks:
  data/human_validation_annotator1.xlsx
  data/human_validation_annotator2.xlsx

Each workbook contains one row per (problem, paraphrase variant) pair.
The annotator fills in a YES/NO judgment for each paraphrase.

Workbook structure
------------------
  Sheet: Instructions
    -- Full annotation guidelines with examples

  Sheet: Annotations
    Columns:
      problem_id      | read-only
      domain          | read-only
      difficulty      | read-only
      source          | read-only
      paraphrase_id   | read-only
      strategy        | read-only
      original_text   | read-only  (P0, the original problem)
      paraphrase_text | read-only  (the variant to judge)
      answer          | read-only  (correct answer, for reference)
      VALID (YES/NO)  | annotator fills this in
      NOTES           | optional free-text comment

After annotation, save the file and pass both annotator files to:
  python 4_consolidate_validation.py

Requirements
------------
  pip install openpyxl

Usage
-----
  python 2_generate_validation_template.py
"""

import json
import os
import sys

import openpyxl
from openpyxl.styles import (
    Alignment, Font, PatternFill, Protection
)
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

DATA_DIR       = os.path.join(os.path.dirname(os.path.abspath(__file__)), "..", "data")
PARAPHRASE_FILE = os.path.join(DATA_DIR, "paraphrases.json")

STRATEGY_DESCRIPTIONS = {
    "P0": "Original (no change)",
    "P1": "Formal restatement -- precise mathematical/logical language",
    "P2": "Informal restatement -- casual everyday language",
    "P3": "Passive restructuring -- passive voice, different sentence order",
    "P4": "Decomposed -- problem split into numbered sub-steps",
    "P5": "Analogical -- same structure, different real-world context",
}

INSTRUCTIONS = """
PARAPHRASE VALIDATION INSTRUCTIONS
===================================

Your task is to judge whether each paraphrase is SEMANTICALLY EQUIVALENT
to the original problem (P0).

A paraphrase is VALID (YES) if:
  1. It asks for exactly the same answer as the original.
  2. All numerical values and constraints are preserved (for math problems).
  3. The answer choices and correct answer letter are preserved (for logic problems).
  4. It does not add new information, hints, or sub-questions not in the original.
  5. It does not remove any information needed to solve the problem.

A paraphrase is INVALID (NO) if:
  1. The correct answer would change.
  2. A number, constraint, or answer choice has been altered.
  3. The problem has been made easier or harder by adding/removing information.
  4. The paraphrase is incomprehensible or grammatically broken.

IMPORTANT:
  - P0 rows are the originals. Mark them YES always (they are included for reference).
  - If you are unsure, mark NO and add a note in the NOTES column.
  - Do not look up answers. Judge based on structure and content only.
  - Work independently. Do not discuss with the other annotator until both are done.

Fill in the VALID (YES/NO) column for every row.
Save the file when done and return it to the experiment coordinator.
"""


def create_workbook(problems: list[dict], annotator_name: str) -> openpyxl.Workbook:
    wb = openpyxl.Workbook()

    # ---- Sheet 1: Instructions ----
    ws_inst = wb.active
    ws_inst.title = "Instructions"
    ws_inst["A1"] = "ANNOTATION INSTRUCTIONS"
    ws_inst["A1"].font = Font(bold=True, size=14)
    ws_inst["A2"] = INSTRUCTIONS
    ws_inst["A2"].alignment = Alignment(wrap_text=True, vertical="top")
    ws_inst.column_dimensions["A"].width = 100
    ws_inst.row_dimensions[2].height = 400

    # ---- Sheet 2: Annotations ----
    ws = wb.create_sheet("Annotations")

    header_fill   = PatternFill("solid", fgColor="2F5496")
    locked_fill   = PatternFill("solid", fgColor="F2F2F2")
    input_fill    = PatternFill("solid", fgColor="FFFACD")
    p0_fill       = PatternFill("solid", fgColor="E8F4E8")

    headers = [
        "problem_id", "domain", "difficulty", "source",
        "paraphrase_id", "strategy_description",
        "original_text (P0)", "paraphrase_text",
        "correct_answer",
        "VALID (YES/NO)",
        "NOTES (optional)",
    ]

    # Write header row
    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=h)
        cell.font      = Font(bold=True, color="FFFFFF", size=11)
        cell.fill      = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    ws.row_dimensions[1].height = 30

    # YES/NO dropdown validation
    dv = DataValidation(
        type="list",
        formula1='"YES,NO"',
        allow_blank=False,
        showDropDown=False,
        showErrorMessage=True,
        errorTitle="Invalid input",
        error="Please select YES or NO from the dropdown.",
    )
    ws.add_data_validation(dv)

    row_idx = 2
    for problem in problems:
        # Find P0 text
        p0_text = next(
            (v["text"] for v in problem["paraphrases"] if v["id"] == "P0"),
            problem.get("question", ""),
        )
        for variant in problem["paraphrases"]:
            is_original = variant["id"] == "P0"
            row_fill    = p0_fill if is_original else locked_fill

            values = [
                problem["id"],
                problem["domain"],
                problem["difficulty"],
                problem.get("source", ""),
                variant["id"],
                STRATEGY_DESCRIPTIONS.get(variant["id"], variant.get("strategy", "")),
                p0_text,
                variant["text"],
                problem["answer"],
                "YES" if is_original else "",   # P0 pre-filled YES
                "",
            ]

            for col_idx, val in enumerate(values, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=val)
                cell.alignment = Alignment(wrap_text=True, vertical="top")
                # Input columns: VALID and NOTES
                if col_idx == 10:
                    cell.fill = PatternFill("solid", fgColor="FFFACD") if not is_original else p0_fill
                    if not is_original:
                        dv.add(cell)
                elif col_idx == 11:
                    cell.fill = PatternFill("solid", fgColor="FFFFFF")
                else:
                    cell.fill = row_fill

            row_idx += 1

    # Column widths
    col_widths = [18, 8, 10, 14, 12, 35, 55, 55, 12, 14, 25]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # Freeze header row
    ws.freeze_panes = "A2"

    # Add annotator name in a prominent cell
    ws_inst["A4"] = f"Annotator: {annotator_name}"
    ws_inst["A4"].font = Font(bold=True, size=12, color="C00000")

    return wb


def main():
    print("=" * 60)
    print("ARCG Step 2: Generate Human Validation Templates")
    print("=" * 60)

    if not os.path.exists(PARAPHRASE_FILE):
        print(f"ERROR: {PARAPHRASE_FILE} not found.")
        print("Run 1_build_and_paraphrase.py first.")
        sys.exit(1)

    with open(PARAPHRASE_FILE) as f:
        problems = json.load(f)

    print(f"Loaded {len(problems)} problems with paraphrases.")

    total_rows = sum(len(p["paraphrases"]) for p in problems)
    print(f"Total annotation rows per annotator: {total_rows}")

    os.makedirs(DATA_DIR, exist_ok=True)

    for annotator_id in [1, 2]:
        name     = f"Annotator {annotator_id}"
        out_path = os.path.join(DATA_DIR, f"human_validation_annotator{annotator_id}.xlsx")
        wb       = create_workbook(problems, name)
        wb.save(out_path)
        print(f"  Saved: {out_path}")

    print(
        "\nInstructions:\n"
        "  1. Send each Excel file to a different human annotator.\n"
        "  2. Annotators work INDEPENDENTLY -- no discussion until both are done.\n"
        "  3. Each annotator fills the 'VALID (YES/NO)' column for every non-P0 row.\n"
        "  4. Collect the completed files and run:\n"
        "       python 4_consolidate_validation.py\n"
    )


if __name__ == "__main__":
    main()
